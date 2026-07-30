#!/usr/bin/env python3
"""Rebuild the GEMI simple model with the broken cells repaired.

The input workbook (GEMI_simple_model_v1_original.xlsx, the Artemis-produced
file) is left untouched. This script writes GEMI_simple_model_v2.xlsx with:

  * The rows that were literal "—" strings turned into real formulas. In v1
    Sales, Transaction revenue, both take rates and Prediction % of Sales were
    placeholder text, so every downstream cell that referenced them evaluated
    to #VALUE! — including EV/Sales, Opex %, EBIT, EBITDA and every CAGR.
  * A real EV bridge off the Q1 2026 10-Q balance sheet. v1 carried
    "Net Cash" = -252.8, which is the related-party loans line on its own; it
    ignored $215.6M of cash, $75.3M of third-party loans, $140.5M of funding
    debt and $18.8M of leases.
  * Market cap and net debt populated for the valuation base year, so the
    revenue scenarios can compute upside. In v1 the base-year EV cell was
    empty, so every "Ups/Downs" in the revenue table was #DIV/0! — the single
    most important output of the model was missing.
  * The EPS / PE scenario table replaced. Multiplying a negative 2030 EPS by a
    zero PE to get a $0.00 target and calling it -100% is a spreadsheet
    artifact, not a valuation. Equity value now comes off EV/Sales with a
    tangible-book floor.
  * A prediction-volume sheet built from Gemini Titan's published daily
    series, reconciling the two-sided convention Artemis uses against the
    one-sided public endpoint, plus the July window analysis.

Values are cached by recalculating through LibreOffice so the file is readable
without Excel and reviewable on GitHub.

Usage: python3 build_simple_model_v2.py
"""

import json
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
V1_SRC = Path.home() / "Downloads" / "gemi-onchain-earnings-model-main" / "GEMI_simple_model.xlsx"
V1_KEEP = ROOT / "GEMI_simple_model_v1_original.xlsx"
V2 = ROOT / "GEMI_simple_model_v2.xlsx"
SERIES = ROOT / "data" / "prediction_volume_series.json"
TABLES_MD = ROOT / "SIMPLE_MODEL_V2_TABLES.md"

PRICE = 4.065
SHARES_OUT = 126.2  # 513 / 4.065, consistent with the v1 market cap input

# Q1 2026 10-Q condensed consolidated balance sheet, USD millions.
# https://www.sec.gov/Archives/edgar/data/2055592/000205559226000050/R2.htm
BS = {
    "cash": 215.623,
    "restricted_cash": 103.747,
    "customer_custodial_funds": 483.774,
    "crypto_assets_held": 271.956,
    "credit_card_receivables_pledged": 183.965,
    "custodial_funds_due_to_customers": 483.656,
    "third_party_loans": 75.349,
    "related_party_loans": 252.574,
    "funding_debt": 140.457,
    "lease_liabilities": 18.764,
    "total_liabilities": 1065.678,
    "total_assets": 1521.816,
    "stockholders_equity": 456.138,
    "intangibles": 134.231,
}

YEARS = [2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030]
# Columns B..I on the Model sheet, matching v1's layout.
YCOL = {y: get_column_letter(2 + i) for i, y in enumerate(YEARS)}

# Operating drivers, carried over from v1 unchanged so the diff is purely
# structural. 2023 and 2024 were blank in v1 and stay blank.
DRIVERS = {
    "exchange_volume_b": {2025: 52.7, 2026: 16.8, 2027: 24, 2028: 30, 2029: 34, 2030: 38},
    "prediction_notional_m": {2025: 2.8, 2026: 745, 2027: 1150, 2028: 1700, 2029: 2200, 2030: 2700},
    "card_volume_b": {2026: 4.3, 2027: 5.5, 2028: 6.8, 2029: 8.0, 2030: 9.2},
    "mtu_k": {2025: 601, 2026: 600, 2027: 660, 2028: 730, 2029: 800, 2030: 870},
    "exchange_revenue": {2025: 93.426, 2026: 47.8, 2027: 64.8, 2028: 78, 2029: 85, 2030: 91.2},
    "otc_revenue": {2025: 4.013, 2026: 17.6, 2027: 18, 2028: 21, 2029: 24, 2030: 27},
    "prediction_revenue": {2025: 0.0, 2026: 2.5, 2027: 3.6, 2028: 5.1, 2029: 6.4, 2030: 7.6},
    "other_transaction_revenue": {2025: 0.581, 2026: 0.8, 2027: 0.9, 2028: 1.0, 2029: 1.1, 2030: 1.2},
    "service_revenue": {2025: 64.639, 2026: 88.2, 2027: 105, 2028: 124, 2029: 148, 2030: 178},
    "other_income": {2025: 16.913, 2026: 17.2, 2027: 18, 2028: 21, 2029: 25, 2030: 30},
    "opex": {2025: 525.228, 2026: 508, 2027: 495, 2028: 490, 2029: 495, 2030: 500},
    "depreciation": {2025: 30.723, 2026: 30, 2027: 32, 2028: 33, 2029: 34, 2030: 35},
    "sbc": {2025: 84.956, 2026: 88, 2027: 85, 2028: 82, 2029: 80, 2030: 78},
    "net_income": {2025: -582.813, 2026: -367, 2027: -310, 2028: -265, 2029: -230, 2030: -190},
    "eps": {2025: -15.52, 2026: -2.97, 2027: -2.42, 2028: -2.02, 2029: -1.72, 2030: -1.39},
}

CONSENSUS = {
    "fy2026_revenue": 193.18,
    "fy2026_eps": -2.897,
    "fy2030_revenue": 394.15,
    "fy2030_eps": -1.775,
    "q2_2026_revenue": 42.84,
    "q2_2026_eps": -0.690,
    "q2_2026_exchange": 12.36,
}

# ---------------------------------------------------------------- styling

H1 = Font(bold=True, size=13)
HDR = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="1F3864")
SECT = Font(bold=True, color="1F3864")
SECT_FILL = PatternFill("solid", fgColor="D9E2F3")
BOLD = Font(bold=True)
MUTED = Font(italic=True, color="666666")
FIX = PatternFill("solid", fgColor="E2EFDA")   # repaired cell
NEWF = PatternFill("solid", fgColor="FFF2CC")  # new input
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM1 = "#,##0.0"
NUM2 = "#,##0.00"
NUM0 = "#,##0"
PCT1 = "0.0%"
BPS = '#,##0.0" bps"'
USD2 = '$#,##0.00'
USDM = '$#,##0.0,,"M"'


def hdr_row(ws, row, labels, start=1):
    for i, label in enumerate(labels):
        c = ws.cell(row, start + i, label)
        c.font = HDR
        c.fill = HDR_FILL
        c.border = BOX
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def section(ws, row, text, span=8):
    c = ws.cell(row, 1, text)
    c.font = SECT
    for i in range(span):
        ws.cell(row, 1 + i).fill = SECT_FILL


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------- Model

def build_model(ws):
    ws["A1"] = f"Gemini Space Station, Inc. (GEMI) — ${PRICE:.3f} — repaired model (v2)"
    ws["A1"].font = H1
    ws["A2"] = (
        "Green cells were placeholder text in v1 and are now formulas. Yellow cells are new inputs. "
        "Operating drivers are unchanged from v1 so the only differences are structural."
    )
    ws["A2"].font = MUTED

    r = 4
    for i, y in enumerate(YEARS):
        c = ws.cell(r, 2 + i, y)
        c.font = HDR
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")
        c.number_format = "0"  # a year is not a quantity; no thousands separator
    ws.cell(r, 1, "USD millions unless noted").font = BOLD

    rows = {}

    def put(row, label, key=None, fmt=NUM1, indent=False):
        c = ws.cell(row, 1, ("  " if indent else "") + label)
        if not indent:
            c.font = BOLD
        if key:
            for y in YEARS:
                v = DRIVERS[key].get(y)
                if v is not None:
                    cell = ws.cell(row, 2 + YEARS.index(y), v)
                    cell.number_format = fmt
        return row

    def formula_row(row, label, tmpl, fmt=NUM1, indent=True, fix=True, years=None):
        c = ws.cell(row, 1, ("  " if indent else "") + label)
        if not indent:
            c.font = BOLD
        for y in years or YEARS:
            col = get_column_letter(2 + YEARS.index(y))
            cell = ws.cell(row, 2 + YEARS.index(y), tmpl.format(c=col))
            cell.number_format = fmt
            if fix:
                cell.fill = FIX
        return row

    section(ws, r + 1, "Operating drivers", 9)
    rows["exch_vol"] = put(r + 2, "Exchange volume ($B)", "exchange_volume_b")
    # Take rate in bps: revenue in $M over volume in $B is 10x bps.
    rows["exch_tr"] = formula_row(
        r + 3, "Exchange take rate", "=IF(N({c}%d)=0,\"\",{c}%d/{c}%d*10)" % (r + 2, r + 9, r + 2), BPS
    )
    rows["pred_not"] = put(r + 4, "Prediction notional ($M, two-sided)", "prediction_notional_m")
    rows["pred_tr"] = formula_row(
        r + 5, "Prediction take rate", "=IF(N({c}%d)=0,\"\",{c}%d/{c}%d*10000)" % (r + 4, r + 11, r + 4), BPS
    )
    rows["card"] = put(r + 6, "Card volume ($B)", "card_volume_b")
    rows["mtu"] = put(r + 7, "Monthly transacting users (k)", "mtu_k", NUM0)

    section(ws, r + 8, "Revenue", 9)
    # shift: revenue block starts at r+9
    rows["exch_rev"] = put(r + 9, "Exchange revenue", "exchange_revenue")
    rows["otc"] = put(r + 10, "OTC revenue", "otc_revenue")
    rows["pred_rev"] = put(r + 11, "Prediction revenue", "prediction_revenue")
    rows["other_txn"] = put(r + 12, "Other transaction revenue", "other_transaction_revenue")
    rows["txn"] = formula_row(
        r + 13, "Transaction revenue",
        "=IF(COUNT({c}%d:{c}%d)=0,\"\",SUM({c}%d:{c}%d))" % (r + 9, r + 12, r + 9, r + 12),
        indent=False,
    )
    rows["svc"] = put(r + 14, "Service revenue", "service_revenue")
    rows["oi"] = put(r + 15, "Other income", "other_income")
    rows["sales"] = formula_row(
        r + 16, "Sales",
        "=IF(COUNT({c}%d,{c}%d,{c}%d)=0,\"\",{c}%d+{c}%d+{c}%d)" % (r + 13, r + 14, r + 15, r + 13, r + 14, r + 15),
        indent=False,
    )
    ws.cell(r + 17, 1, "  Sales Y/Y")  # populated below, needs a prior-column reference
    rows["sales_yy"] = r + 17
    rows["pred_pct"] = formula_row(
        r + 18, "Prediction % of Sales",
        "=IFERROR(IF(N({c}%d)=0,\"\",{c}%d/{c}%d),\"\")" % (r + 16, r + 11, r + 16), PCT1
    )

    # Sales Y/Y needs a previous-column reference; patch it manually.
    for y in YEARS:
        i = YEARS.index(y)
        col = get_column_letter(2 + i)
        cell = ws.cell(r + 17, 2 + i)
        if i == 0:
            cell.value = ""
            continue
        prev = get_column_letter(1 + i)
        cell.value = (
            f'=IFERROR(IF(OR(N({prev}{r+16})=0,N({col}{r+16})=0),"",'
            f'{col}{r+16}/{prev}{r+16}-1),"")'
        )
        cell.number_format = PCT1
        cell.fill = FIX

    section(ws, r + 19, "Costs and earnings", 9)
    rows["opex"] = put(r + 20, "Opex", "opex")
    rows["opex_pct"] = formula_row(
        r + 21, "Opex %", "=IFERROR(IF(N({c}%d)=0,\"\",{c}%d/{c}%d),\"\")" % (r + 16, r + 20, r + 16), PCT1
    )
    rows["ebit"] = formula_row(
        r + 22, "EBIT", "=IF(N({c}%d)=0,\"\",{c}%d-{c}%d)" % (r + 16, r + 16, r + 20), indent=False
    )
    rows["ebit_pct"] = formula_row(
        r + 23, "EBIT %", "=IFERROR(IF(N({c}%d)=0,\"\",{c}%d/{c}%d),\"\")" % (r + 16, r + 22, r + 16), PCT1
    )
    rows["dep"] = put(r + 24, "Depreciation", "depreciation")
    rows["ebitda"] = formula_row(
        r + 25, "EBITDA", "=IF(N({c}%d)=0,\"\",{c}%d+{c}%d)" % (r + 22, r + 22, r + 24), indent=False
    )
    rows["sbc"] = put(r + 26, "SBC", "sbc")
    rows["ebitda_sbc"] = formula_row(
        r + 27, "EBITDA ex SBC", "=IF(N({c}%d)=0,\"\",{c}%d+{c}%d)" % (r + 25, r + 25, r + 26)
    )
    rows["ni"] = put(r + 28, "Net Income", "net_income")
    rows["eps"] = put(r + 29, "EPS ($)", "eps", NUM2)
    rows["shares"] = formula_row(
        r + 30, "Implied diluted shares (M)",
        "=IFERROR(IF(N({c}%d)=0,\"\",{c}%d/{c}%d),\"\")" % (r + 29, r + 28, r + 29), NUM1
    )
    ws.cell(r + 31, 1, (
        "Implied share count is a v2 addition. It shows the model already assumes roughly 124M shares in 2026 "
        "rising to 137M by 2030, consistent with 126.2M outstanding today. FY2025's 37.6M is a pre-IPO weighted "
        "average, so any growth rate computed off a 2025 per-share figure is meaningless."
    )).font = MUTED

    widths(ws, {"A": 34, **{get_column_letter(2 + i): 11 for i in range(8)}})
    return rows


# ---------------------------------------------------------------- EV bridge

def build_ev(ws):
    ws["A1"] = "EV bridge — Q1 2026 10-Q balance sheet"
    ws["A1"].font = H1
    ws["A2"] = (
        "v1 used a single input, Net Cash = -252.8, which equals the related-party loans line alone. "
        "It netted no cash and omitted three other debt lines. The variants below are all defensible; "
        "variant 3 feeds the scenarios because it excludes the two matched books that are not shareholder claims."
    )
    ws["A2"].font = MUTED

    r = 4
    section(ws, r, "Balance sheet, USD millions (Mar 31, 2026)", 4)
    hdr_row(ws, r + 1, ["Line", "Amount", "Treatment", "Why"])
    lines = [
        ("Cash and cash equivalents", BS["cash"], "Corporate cash", "Unencumbered. Nets against debt."),
        ("Restricted cash", BS["restricted_cash"], "Excluded", "Not available to service claims or fund operations."),
        ("Customer custodial funds", BS["customer_custodial_funds"], "Excluded",
         f"Offset by custodial funds due to customers of ${BS['custodial_funds_due_to_customers']:.1f}M. Customer money, not shareholder value."),
        ("Crypto assets held", BS["crypto_assets_held"], "Variant 4 only",
         "Partly financed by the crypto loan lines, so crediting it in full while also counting those loans double counts."),
        ("Credit card receivables pledged", BS["credit_card_receivables_pledged"], "Excluded with funding debt",
         "Matched book: pledged against the funding debt below."),
        ("Third party loans", -BS["third_party_loans"], "Debt", "Includes the bitcoin repurchase agreement."),
        ("Related party loans", -BS["related_party_loans"], "Debt", "The only line v1 captured."),
        ("Funding debt", -BS["funding_debt"], "Excluded with pledged receivables", "Secured by the credit card receivables."),
        ("Lease liabilities", -BS["lease_liabilities"], "Debt", "Non-current."),
    ]
    for i, (label, amt, treat, why) in enumerate(lines):
        ws.cell(r + 2 + i, 1, label)
        c = ws.cell(r + 2 + i, 2, amt)
        c.number_format = NUM1
        ws.cell(r + 2 + i, 3, treat)
        ws.cell(r + 2 + i, 4, why).alignment = Alignment(wrap_text=True, vertical="top")

    b = r + 2 + len(lines) + 1
    # Debt lines are the negative entries in `lines`; derive their rows so the
    # total cannot drift if a line is inserted above.
    debt_rows = [r + 2 + i for i, (_, amt, _, _) in enumerate(lines) if amt < 0]
    ws.cell(b, 1, "Total debt (all four lines)").font = BOLD
    ws.cell(b, 2, "=-(" + "+".join(f"B{x}" for x in debt_rows) + ")").number_format = NUM1
    ws.cell(b, 2).fill = NEWF

    r2 = b + 2
    section(ws, r2, "EV variants", 5)
    hdr_row(ws, r2 + 1, ["#", "Definition", "Net debt", "EV", "Note"])
    variants = [
        ("1", "v1 as delivered: cap + related-party loans only", BS["related_party_loans"],
         "What the v1 scenarios were valued against."),
        ("2", "cap + all debt - corporate cash", BS["third_party_loans"] + BS["related_party_loans"]
         + BS["funding_debt"] + BS["lease_liabilities"] - BS["cash"], "Most literal reading."),
        ("3", "variant 2, excluding the matched credit-card book", BS["third_party_loans"]
         + BS["related_party_loans"] + BS["lease_liabilities"] - BS["cash"],
         "Selected. Funding debt and pledged receivables offset each other."),
        ("4", "variant 3, also crediting crypto assets held", BS["third_party_loans"]
         + BS["related_party_loans"] + BS["lease_liabilities"] - BS["cash"] - BS["crypto_assets_held"],
         "Most generous. Risks double counting against the crypto loans."),
    ]
    for i, (num, defn, nd, note) in enumerate(variants):
        row = r2 + 2 + i
        ws.cell(row, 1, num)
        ws.cell(row, 2, defn)
        ws.cell(row, 3, nd).number_format = NUM1
        ws.cell(row, 4, f"=513+C{row}").number_format = NUM1
        ws.cell(row, 5, note).alignment = Alignment(wrap_text=True, vertical="top")
        if num == "3":
            for cc in range(1, 6):
                ws.cell(row, cc).font = BOLD

    sel = r2 + 4  # variant 3 row
    r3 = r2 + 2 + len(variants) + 1
    ws.cell(r3, 1, "Market cap").font = BOLD
    ws.cell(r3, 2, 513).number_format = NUM1
    ws.cell(r3, 2).fill = NEWF
    ws.cell(r3 + 1, 1, "Selected EV (variant 3)").font = BOLD
    ws.cell(r3 + 1, 2, f"=D{sel}").number_format = NUM1
    ws.cell(r3 + 1, 2).fill = NEWF
    ws.cell(r3 + 2, 1, "Tangible book equity").font = BOLD
    ws.cell(r3 + 2, 2, BS["stockholders_equity"] - BS["intangibles"]).number_format = NUM1
    ws.cell(r3 + 3, 1, "Tangible book per share ($)").font = BOLD
    ws.cell(r3 + 3, 2, f"=B{r3+2}/{SHARES_OUT}").number_format = USD2

    r4 = r3 + 5
    section(ws, r4, "The $803M cash claim", 5)
    ws.cell(r4 + 1, 1, (
        f"${BS['cash']:.1f}M cash + ${BS['restricted_cash']:.1f}M restricted + "
        f"${BS['customer_custodial_funds']:.1f}M customer custodial funds = "
        f"${BS['cash'] + BS['restricted_cash'] + BS['customer_custodial_funds']:.1f}M. "
        f"The custodial component is offset almost exactly by ${BS['custodial_funds_due_to_customers']:.1f}M of "
        "custodial funds due to customers, and restricted cash is not available by definition. Unencumbered "
        f"corporate cash is ${BS['cash']:.1f}M. An asset-support argument built on the $803M figure counts "
        "customer money as shareholder value."
    )).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r4 + 1, start_column=1, end_row=r4 + 3, end_column=5)

    ws.cell(r4 + 5, 1, (
        "Separately, the $372M crypto figure does not tie to the last filed balance sheet, which shows "
        f"${BS['crypto_assets_held']:.1f}M of crypto assets held at Mar 31, 2026 and ${439.622:.1f}M at Dec 31, 2025. "
        "Treat $372M as unverified until the Q2 10-Q lands."
    )).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r4 + 5, start_column=1, end_row=r4 + 6, end_column=5)

    widths(ws, {"A": 42, "B": 13, "C": 13, "D": 13, "E": 46})
    return {"ev_cell": f"'EV Bridge'!$B${r3+1}", "tbv_ps": f"'EV Bridge'!$B${r3+3}"}


# ---------------------------------------------------------------- scenarios

def build_scenarios(ws, refs, model_rows):
    ws["A1"] = "Scenarios — 2030 estimates"
    ws["A1"].font = H1
    ws["A2"] = (
        "v1's revenue table could not compute upside because the base-year EV cell was blank, and its EPS table "
        "multiplied a negative EPS by a zero PE to produce a $0.00 target and a -100% return. Both are fixed here."
    )
    ws["A2"].font = MUTED

    ev = refs["ev_cell"]
    tbv = refs["tbv_ps"]

    r = 4
    section(ws, r, "EV / Sales on 2030 revenue", 7)
    hdr_row(ws, r + 1, ["Scenario", "2030 revenue", "EV/Sales", "Implied EV",
                        "Implied equity", "Implied price", "Upside"])
    scen = [("Bull", 470, 5.0), ("Base", 335, 3.0), ("Bear", 210, 1.75)]
    for i, (name, rev, mult) in enumerate(scen):
        row = r + 2 + i
        ws.cell(row, 1, name).font = BOLD
        ws.cell(row, 2, rev).number_format = NUM1
        ws.cell(row, 3, mult).number_format = NUM2
        ws.cell(row, 4, f"=B{row}*C{row}").number_format = NUM1
        # Equity = EV - net debt; net debt = EV - market cap at the selected variant.
        ws.cell(row, 5, f"=D{row}-({ev}-513)").number_format = NUM1
        ws.cell(row, 6, f"=MAX(E{row}/{SHARES_OUT},{tbv})").number_format = USD2
        ws.cell(row, 7, f"=F{row}/{PRICE}-1").number_format = PCT1
        for cc in range(4, 8):
            ws.cell(row, cc).fill = FIX

    n = r + 2 + len(scen)
    ws.cell(n + 1, 1, (
        f"Implied price floors at tangible book per share, so a bear case does not print a value below the "
        f"company's hard assets. Upside is measured against the ${PRICE:.3f} close."
    )).font = MUTED

    r2 = n + 3
    section(ws, r2, "Why the EPS / PE table was removed", 7)
    ws.cell(r2 + 1, 1, (
        "v1 valued Base and Bear at EPS x PE = -1.39 x 0 = $0.00 and -2.60 x 0 = $0.00, reporting -100.0% for "
        "both. A PE multiple cannot value negative earnings, and a business carrying "
        f"${BS['stockholders_equity']:.1f}M of book equity and ${BS['cash']:.1f}M of corporate cash is not worth zero. "
        "The Bull row was the only one doing work, and even there $0.20 x 35 = $7.00 is a PE applied to the first "
        "profitable year with no discounting. Sales-based valuation with a book floor is the defensible frame while "
        "the company is loss-making."
    )).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r2 + 1, start_column=1, end_row=r2 + 4, end_column=7)

    r3 = r2 + 6
    section(ws, r3, "Model versus consensus", 7)
    hdr_row(ws, r3 + 1, ["Metric", "Model", "Consensus", "Gap", "", "", ""])
    comp = [
        ("FY2026 revenue", "=Model!E20", CONSENSUS["fy2026_revenue"]),
        ("FY2026 EPS", "=Model!E33", CONSENSUS["fy2026_eps"]),
        ("FY2030 revenue", "=Model!I20", CONSENSUS["fy2030_revenue"]),
        ("FY2030 EPS", "=Model!I33", CONSENSUS["fy2030_eps"]),
    ]
    for i, (label, mv, cv) in enumerate(comp):
        row = r3 + 2 + i
        ws.cell(row, 1, label)
        ws.cell(row, 2, mv).number_format = NUM2
        ws.cell(row, 3, cv).number_format = NUM2
        ws.cell(row, 4, f"=B{row}-C{row}").number_format = NUM2
    ws.cell(r3 + 2 + len(comp) + 1, 1, (
        "The model sits below consensus on revenue in both years but above it on 2030 EPS. That combination "
        "requires materially better margins than the street assumes on a smaller revenue base, which is an "
        "implicit call the v1 file never stated. Either the opex path is too optimistic or the revenue path is "
        "too conservative."
    )).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r3 + 2 + len(comp) + 1, start_column=1, end_row=r3 + 2 + len(comp) + 3, end_column=7)

    widths(ws, {"A": 30, "B": 14, "C": 14, "D": 14, "E": 15, "F": 14, "G": 11})


# ---------------------------------------------------------------- prediction volume

def build_prediction(ws):
    ws["A1"] = "Prediction volume — Gemini Titan published daily series"
    ws["A1"].font = H1
    ws["A2"] = (
        "Source: api.gemini.com/v1/prediction-markets/volume/{date}, the public surface of Titan's DCM daily "
        "publication obligation. Full daily coverage since 2025-12-15."
    )
    ws["A2"].font = MUTED

    data = json.loads(SERIES.read_text())
    daily = {d["date"]: sum(d["byCategory"].values()) for d in data["daily"] if d["status"] == "ok"}

    def total(a, b):
        return sum(v for k, v in daily.items() if a <= k <= b)

    def per_day(a, b):
        ks = [k for k in daily if a <= k <= b]
        return (total(a, b) / len(ks), len(ks)) if ks else (0, 0)

    r = 4
    section(ws, r, "Convention reconciliation against the Artemis warehouse series", 6)
    hdr_row(ws, r + 1, ["Window", "Public endpoint (one-sided)", "Artemis warehouse",
                        "Ratio", "Children rows", "Read"])
    recon = [
        ("2025-12-15 to 2025-12-31", "2025-12-15", "2025-12-31", 2_774_094),
        ("Q1 2026", "2026-01-01", "2026-03-31", 119_336_290),
        ("Q2 2026", "2026-04-01", "2026-06-30", 229_207_575),
    ]
    for i, (label, a, b, art) in enumerate(recon):
        row = r + 2 + i
        mine = total(a, b)
        kids = sum(sum(d["bySubcategory"].values()) for d in data["daily"]
                   if d["status"] == "ok" and a <= d["date"] <= b)
        ws.cell(row, 1, label)
        ws.cell(row, 2, mine).number_format = NUM0
        ws.cell(row, 3, art).number_format = NUM0
        ws.cell(row, 4, art / mine).number_format = NUM2
        ws.cell(row, 5, kids).number_format = NUM0
        ws.cell(row, 6, "Exactly 2x" if abs(art / mine - 2) < 0.002 else "Near 2x")
    ws.cell(r + 2 + len(recon) + 1, 1, (
        "The December window contains only one category row and no children, yet Artemis is still exactly 2.00x. "
        "That rules out hierarchy double counting and points to a two-sided convention: Artemis counts both legs "
        "of each match, the public endpoint counts one. Revenue is unaffected because the take rate is calibrated "
        "and applied on the same basis, but any cross-venue share comparison must put all venues on one convention."
    )).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r + 2 + len(recon) + 1, start_column=1, end_row=r + 2 + len(recon) + 3, end_column=6)

    r2 = r + 2 + len(recon) + 5
    section(ws, r2, "Half-month run rates — the -26% claim", 5)
    hdr_row(ws, r2 + 1, ["Window", "Per day (one-sided)", "Days", "vs prior window", ""])
    wins = [
        ("Q2 2026 average", "2026-04-01", "2026-06-30"),
        ("Jun 1-15", "2026-06-01", "2026-06-15"),
        ("Jun 16-30", "2026-06-16", "2026-06-30"),
        ("Jul 1-15", "2026-07-01", "2026-07-15"),
        ("Jul 16-28", "2026-07-16", "2026-07-28"),
    ]
    vals = {}
    for i, (label, a, b) in enumerate(wins):
        row = r2 + 2 + i
        pdv, n = per_day(a, b)
        vals[label] = pdv
        ws.cell(row, 1, label)
        ws.cell(row, 2, pdv).number_format = NUM0
        ws.cell(row, 3, n)
    r3 = r2 + 2 + len(wins) + 1
    section(ws, r3, "Comparisons", 5)
    hdr_row(ws, r3 + 1, ["Comparison", "Change", "", "", ""])
    comps = [
        ("Jul 1-15 vs Jun 1-15", vals["Jul 1-15"] / vals["Jun 1-15"] - 1),
        ("Jul 1-15 vs Jun 16-30", vals["Jul 1-15"] / vals["Jun 16-30"] - 1),
        ("Jul 1-15 vs Q2 average", vals["Jul 1-15"] / vals["Q2 2026 average"] - 1),
        ("Jul 16-28 vs Jul 1-15", vals["Jul 16-28"] / vals["Jul 1-15"] - 1),
        ("Jul 16-28 vs Jun 16-30", vals["Jul 16-28"] / vals["Jun 16-30"] - 1),
    ]
    for i, (label, ch) in enumerate(comps):
        row = r3 + 2 + i
        ws.cell(row, 1, label)
        c = ws.cell(row, 2, ch)
        c.number_format = PCT1
        if abs(ch + 0.263) < 0.005:
            c.font = BOLD
            ws.cell(row, 3, "matches the quoted -26%")
    ws.cell(r3 + 2 + len(comps) + 1, 1, (
        "The -26% figure reconciles to Jul 16-28 versus Jun 16-30, the back half of each month, not the first "
        "half of July. Jul 1-15 was up 49% on Jun 1-15 and up 45% on the Q2 average. July as a whole ran the "
        "highest per-day volume of any month since launch. There is a real late-July deceleration and it is worth "
        "tracking, but it is the second half of the month, and the level is still above the Q2 average rather "
        "than below it."
    )).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r3 + 2 + len(comps) + 1, start_column=1, end_row=r3 + 2 + len(comps) + 4, end_column=5)

    r4 = r3 + 2 + len(comps) + 6
    section(ws, r4, "Monthly series", 4)
    hdr_row(ws, r4 + 1, ["Month", "Total (one-sided)", "Per day", "Days"])
    months = [("2026-04", "2026-04-01", "2026-04-30"), ("2026-05", "2026-05-01", "2026-05-31"),
              ("2026-06", "2026-06-01", "2026-06-30"), ("2026-07 (1-28)", "2026-07-01", "2026-07-28")]
    for i, (label, a, b) in enumerate(months):
        row = r4 + 2 + i
        pdv, n = per_day(a, b)
        ws.cell(row, 1, label)
        ws.cell(row, 2, total(a, b)).number_format = NUM0
        ws.cell(row, 3, pdv).number_format = NUM0
        ws.cell(row, 4, n)

    widths(ws, {"A": 30, "B": 24, "C": 16, "D": 14, "E": 26, "F": 16})


# ---------------------------------------------------------------- review sheet

FINDINGS = [
    ("Blocking", "Sales row was placeholder text",
     "Model!B18:I18 held the string \"—\", so Sales never computed. Every cell referencing it errored: EV/Sales, "
     "Opex %, EBIT, EBIT %, EBITDA, Sales Y/Y and all CAGRs.",
     "Fixed. Sales = transaction + service + other income."),
    ("Blocking", "Revenue scenarios had no upside column",
     "Market cap and net cash were populated only for 2026-2030, leaving the base-year EV cell empty, so every "
     "Ups/Downs formula divided by zero. The model's headline output was missing.",
     "Fixed. Scenarios now reference an explicit selected-EV cell. Bull is +207% on EV."),
    ("Blocking", "EPS / PE table is analytically void",
     "Base -1.39 x 0 PE = $0.00 and Bear -2.60 x 0 = $0.00, both reported as -100%. A PE cannot value negative "
     "earnings, and the company holds $456.1M of book equity and $215.6M of corporate cash.",
     "Removed. Replaced with EV/Sales to equity value, floored at tangible book per share."),
    ("Blocking", "Net Cash input is one debt line, not net cash",
     "-252.8 equals the related-party loans line ($252.574M) alone. It nets no cash and omits third-party loans "
     "($75.3M), funding debt ($140.5M) and leases ($18.8M).",
     "Fixed. New EV Bridge sheet builds four variants off the Q1 2026 balance sheet; variant 3 is selected."),
    ("Material", "The $803M cash claim counts customer money",
     "$803.1M = cash $215.6M + restricted $103.7M + customer custodial funds $483.8M. The custodial line is "
     "offset by $483.7M of custodial funds due to customers. Unencumbered corporate cash is $215.6M.",
     "Documented on the EV Bridge sheet. The asset-support argument does not survive as stated."),
    ("Material", "The -26% July prediction-volume claim is the wrong window",
     "-26.3% reconciles to Jul 16-28 versus Jun 16-30. Jul 1-15 was +49.3% on Jun 1-15 and +44.9% on the Q2 "
     "average; July ran the highest per-day volume since launch.",
     "Corrected on the Prediction Volume sheet. The late-July deceleration is real; the label is not."),
    ("Material", "The bps share series cannot be reconciled",
     "Gemini share of three-venue notional is quoted at 36.35 / 20.46 / 9.29 bps for May / Jun / Jul days 1-15. "
     "Holding those shares against actual Gemini volume implies the three-venue denominator grew from $12.5B to "
     "$59.2B in two months, a 4.7x increase, against the note's own claim that Kalshi and Polymarket grew 63%.",
     "Flagged. Either the Gemini numerator is truncated for July or the venues are on mixed conventions. "
     "Needs a rebuild before the share-loss conclusion can be used."),
    ("Material", "Two-sided versus one-sided volume convention is undeclared",
     "The warehouse series is exactly 2.00x the public endpoint, including in a December window with no "
     "subcategory rows, so it counts both legs of each match. Revenue is unaffected, but the 34 bps take rate is "
     "half the one-sided equivalent of 74 bps, which changes how it reads against Kalshi at 102-119 bps.",
     "Documented. Cross-venue comparisons must normalise convention first."),
    ("Material", "Below consensus on revenue, above it on EPS",
     "Model FY2030 revenue of $335.0M is 15.0% below consensus $394.15M, yet model EPS of -$1.39 is $0.39 better "
     "than consensus -$1.775. That is a materially better margin call on a smaller base, never stated.",
     "Surfaced in a comparison block. Reconcile the opex path or the revenue path."),
    ("Minor", "$372M crypto does not tie to the filings",
     "The last filed balance sheet shows $272.0M of crypto assets held at Mar 31, 2026 and $439.6M at Dec 31, 2025.",
     "Flagged as unverified."),
    ("Minor", "Take rate rows were placeholder text",
     "Exchange and prediction take rates were \"—\" throughout, hiding the assumption that exchange take rate "
     "steps from 17.7 bps in FY2025 to 28.5 bps in FY2026 then decays to 24.0 bps by 2030.",
     "Fixed. Both now compute in bps. The FY2026 step is supported by Q1 2026 actuals at 27.3 bps."),
    ("Minor", "CAGR formulas are undefined on negative bases",
     "The EPS CAGR raised the ratio of 2030 EPS to FY2025 EPS to a fractional power across a sign change, and the "
     "revenue CAGR divided by the broken Sales row.",
     "Removed. FY2025 EPS uses a pre-IPO 37.6M weighted share count and is not comparable per share anyway."),
    ("Confirmed", "The spot volume collapse checks out",
     "An independent CoinGecko series gives Gemini spot at $37.5M/day in Q2 and $23.0M/day for Jul 1-28, "
     "-38.7%, against the note's $48.4M and $27.9M, -42%. Levels differ by source, direction and magnitude agree.",
     "No change. This is the strongest part of the analysis and the Q3 exchange-revenue gap is real."),
    ("Confirmed", "Prediction materiality is right",
     "At 34 bps on $229.2M two-sided notional, Q2 prediction revenue is about $0.78M, 1.8% of a $43.8M revenue "
     "estimate and roughly $0.006 of EPS on 126.2M shares. The independent one-sided calibration gives $0.858M.",
     "No change. Both methods agree that the segment is immaterial to Q2."),
]


def build_review(ws):
    ws["A1"] = "Review of the v1 simple model"
    ws["A1"].font = H1
    ws["A2"] = (
        "Blocking means the v1 file does not produce the number it claims to. Material means the number computes "
        "but the conclusion drawn from it does not hold. Confirmed means it was checked against an independent "
        "source and stands."
    )
    ws["A2"].font = MUTED
    hdr_row(ws, 4, ["Severity", "Finding", "Evidence", "Resolution in v2"])
    for i, (sev, name, ev, res) in enumerate(FINDINGS):
        row = 5 + i
        c = ws.cell(row, 1, sev)
        c.font = BOLD
        c.fill = PatternFill("solid", fgColor={
            "Blocking": "F8CBAD", "Material": "FFE699", "Minor": "E2EFDA", "Confirmed": "D9E2F3"}[sev])
        ws.cell(row, 2, name).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 3, ev).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 4, res).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 58
    widths(ws, {"A": 11, "B": 34, "C": 74, "D": 52})


# ---------------------------------------------------------------- sources

def build_sources(ws, v1_sources):
    ws["A1"] = "Sources"
    ws["A1"].font = H1
    hdr_row(ws, 3, ["id", "kind", "url", "excerpt"])
    row = 4
    for rec in v1_sources:
        for j, v in enumerate(rec):
            ws.cell(row, 1 + j, v).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 46
        row += 1
    added = [
        ("src_titan_volume_api", "exchange API",
         "https://api.gemini.com/v1/prediction-markets/volume/{date}",
         "Gemini Titan published daily prediction-market volume by category, one-sided. Full daily coverage from "
         "2025-12-15. Q1 2026 59,667,196; Q2 2026 115,249,040; Jul 1-28 2026 45,646,530."),
        ("src_titan_volume_spec", "docs",
         "https://developer.gemini.com/prediction-markets-spec/volume",
         "Endpoint spec. GET /v1/prediction-markets/volume/{date} and /hourly, unauthenticated, response is an "
         "array of {categoryPath, volume}. categoryPath is hierarchical, so only depth-1 rows may be summed."),
        ("src_titan_rulebook", "rulebook",
         "https://www.cftc.gov/filings/orgrules/rules03252641973.pdf",
         "Gemini Titan LLC rulebook v1.8. Rule 2.17(b): Gemini shall make public on a daily basis information on "
         "settlement prices, volume, open interest and opening and closing ranges for actively traded Contracts."),
        ("src_titan_exhibit_l", "CFTC filing",
         "https://www.cftc.gov/sites/default/files/filings/documents/2025/orgdcmgmniexhibitl250515.pdf",
         "Titan DCM application Exhibit L, Core Principle 8 Daily Publication of Trading Information."),
        ("src_cftc_1601", "regulation",
         "https://www.law.cornell.edu/cfr/text/17/16.01",
         "17 CFR 16.01. Paragraph (e) requires reporting markets to make volume, price and critical date "
         "information readily available to the news media and the general public without charge no later than the "
         "business day following. Transaction-level trade data is 16.02 and is not published."),
        ("src_gemi_bs_q1_2026", "10-Q",
         "https://www.sec.gov/Archives/edgar/data/2055592/000205559226000050/R2.htm",
         "GEMI condensed consolidated balance sheet, Mar 31 2026, USD thousands: cash 215,623; restricted cash "
         "103,747; customer custodial funds 483,774; crypto assets held 271,956; credit card receivables pledged "
         "183,965; custodial funds due to customers 483,656; third party loans 75,349; related party loans "
         "252,574; funding debt 140,457; lease liabilities 18,764; total liabilities 1,065,678; total assets "
         "1,521,816; stockholders equity 456,138; intangibles 134,231."),
        ("src_gemini_pred_fees", "fee schedule",
         "https://www.gemini.com/fees/predictions",
         "Gemini Predictions fees: Fee = Rate x Contracts x Price x (1 - Price), taker rate 0.07, maker rate "
         "0.0175, no settlement fees."),
        ("src_coingecko_gemini_vol", "market data",
         "https://api.coingecko.com/api/v3/exchanges/gemini/volume_chart",
         "Independent Gemini spot exchange volume. Q1 2026 $62.7M/day, Q2 2026 $37.5M/day, Jul 1-28 2026 "
         "$23.0M/day, July versus Q2 -38.7%."),
        ("src_hood_q2_2026", "press release",
         "https://investors.robinhood.com/news-releases",
         "Robinhood Q2 2026, reported 2026-07-29: total net revenues $1.31B, crypto revenue $100M down 38% YoY, "
         "event contracts revenue $156M up over 10x, event contracts traded 13.6B, crypto notional $40B "
         "(app $18B, Bitstamp $22B), net income $573M, diluted EPS $0.62."),
    ]
    for rec in added:
        for j, v in enumerate(rec):
            c = ws.cell(row, 1 + j, v)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.fill = NEWF
        ws.row_dimensions[row].height = 46
        row += 1
    widths(ws, {"A": 26, "B": 15, "C": 60, "D": 96})


# ---------------------------------------------------------------- main

def read_v1_sources():
    wb = load_workbook(V1_KEEP, data_only=True)
    ws = wb["Sources"]
    out = []
    for r in range(2, ws.max_row + 1):
        rec = [ws.cell(r, c).value or "" for c in range(1, 5)]
        if any(rec):
            out.append(rec)
    return out


def recalc(path):
    """Recalculate through LibreOffice so cached values exist for non-Excel readers."""
    outdir = path.parent / "_recalc"
    outdir.mkdir(exist_ok=True)
    try:
        subprocess.run(
            ["soffice", "--headless", "--norestore", "--convert-to", "xlsx",
             "--outdir", str(outdir), str(path)],
            check=True, capture_output=True, timeout=180,
        )
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired, FileNotFoundError) as e:
        print(f"  WARNING: recalculation skipped ({e}). Formulas will have no cached values.")
        shutil.rmtree(outdir, ignore_errors=True)
        return False
    produced = outdir / path.name
    if produced.exists():
        shutil.move(str(produced), str(path))
        shutil.rmtree(outdir, ignore_errors=True)
        return True
    shutil.rmtree(outdir, ignore_errors=True)
    return False


def _fmt(cell):
    """Render a cell the way the spreadsheet would, so percentages read as
    percentages rather than raw decimals."""
    v = cell.value
    if v is None:
        return ""
    if not isinstance(v, (int, float)) or isinstance(v, bool):
        return str(v).replace("|", "\\|").replace("\n", " ")
    nf = cell.number_format or ""
    if "%" in nf:
        return f"{v * 100:,.1f}%"
    if "bps" in nf:
        return f"{v:,.1f} bps"
    if nf.startswith("$"):
        return f"${v:,.2f}"
    if nf == "0":
        return str(int(v))
    if isinstance(v, int) or float(v).is_integer():
        return f"{int(v):,}"
    return f"{v:,.2f}" if abs(v) < 1000 else f"{v:,.1f}"


def dump_tables():
    """Write every sheet's computed values to markdown so the workbook is
    reviewable on GitHub or by a model with no spreadsheet engine."""
    wb = load_workbook(V2, data_only=True)
    lines = [
        "# GEMI Simple Model v2 — computed tables",
        "",
        "Plain-text rendering of every sheet in `GEMI_simple_model_v2.xlsx`, with formulas",
        "resolved to values. Generated by `build_simple_model_v2.py`; do not hand-edit.",
        "This exists so the workbook can be read on GitHub or by an AI assistant without Excel.",
        "",
    ]
    for ws in wb.worksheets:
        lines += [f"## {ws.title}", ""]
        # Collect rows, trimming trailing blanks so single-cell notes stay narrow.
        grid = []
        for r in range(1, ws.max_row + 1):
            vals = [_fmt(ws.cell(r, c)) for c in range(1, ws.max_column + 1)]
            while vals and vals[-1] == "":
                vals.pop()
            grid.append(vals)

        # Emit each run of consecutive non-empty rows as one table, with the
        # separator GitHub needs after the run's first row.
        i = 0
        while i < len(grid):
            if not grid[i]:
                i += 1
                continue
            block = []
            while i < len(grid) and grid[i]:
                block.append(grid[i])
                i += 1
            width = max(len(v) for v in block)
            # A lone one-cell row is prose, not a table; emit it as a paragraph.
            if width == 1 and len(block) == 1:
                lines += [block[0][0], ""]
                continue
            for j, vals in enumerate(block):
                vals = vals + [""] * (width - len(vals))
                lines.append("| " + " | ".join(vals) + " |")
                if j == 0:
                    lines.append("|" + "|".join(["---"] * width) + "|")
            lines.append("")
    TABLES_MD.write_text("\n".join(lines))


def main():
    if not V1_KEEP.exists():
        if not V1_SRC.exists():
            sys.exit(f"Cannot find the original workbook at {V1_SRC} or {V1_KEEP}")
        shutil.copy2(V1_SRC, V1_KEEP)
        print(f"  archived original -> {V1_KEEP.name}")

    v1_sources = read_v1_sources()

    wb = Workbook()
    model = wb.active
    model.title = "Model"
    rows = build_model(model)
    refs = build_ev(wb.create_sheet("EV Bridge"))
    build_scenarios(wb.create_sheet("Scenarios"), refs, rows)
    build_prediction(wb.create_sheet("Prediction Volume"))
    build_review(wb.create_sheet("Review"))
    build_sources(wb.create_sheet("Sources"), v1_sources)
    wb.save(V2)
    print(f"  wrote {V2.name}")

    if recalc(V2):
        print("  recalculated via LibreOffice, cached values embedded")
    dump_tables()
    print(f"  wrote {TABLES_MD.name}")


if __name__ == "__main__":
    main()
